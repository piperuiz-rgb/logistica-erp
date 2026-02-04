import streamlit as st
import pandas as pd
import os
import io
import json
import re
from datetime import datetime
from openpyxl import load_workbook
from streamlit_local_storage import LocalStorage

# =========================================================
# CONFIG
# =========================================================
st.set_page_config(page_title="Peticiones", layout="wide")

LS_KEY = "peticiones_estado_v1"
localS = LocalStorage()

# =========================================================
# LOCAL STORAGE (compatibilidad entre firmas)
# =========================================================
def ls_get(item_key: str, ss_key: str):
    try:
        out = localS.getItem(item_key, key=ss_key)
        if ss_key in st.session_state and st.session_state[ss_key]:
            return st.session_state[ss_key]
        return out
    except TypeError:
        pass

    try:
        out = localS.getItem(item_key, ss_key)
        if ss_key in st.session_state and st.session_state[ss_key]:
            return st.session_state[ss_key]
        return out
    except TypeError:
        pass

    try:
        return localS.getItem(item_key)
    except TypeError:
        return None


def ls_set(item_key: str, value: str) -> None:
    localS.setItem(item_key, value)


# =========================================================
# STATE HELPERS
# =========================================================
def _serialize_state() -> dict:
    return {
        "carrito": st.session_state.get("carrito", {}),
        "fecha_str": st.session_state.get("fecha_str"),
        "origen": st.session_state.get("origen"),
        "destino": st.session_state.get("destino"),
        "ref_peticion": st.session_state.get("ref_peticion", ""),
    }


def _apply_state(payload: dict) -> None:
    if not isinstance(payload, dict):
        return
    st.session_state.carrito = payload.get("carrito", {}) or {}
    for k in ("origen", "destino", "ref_peticion", "fecha_str"):
        if payload.get(k) is not None:
            st.session_state[k] = payload[k]


def mark_dirty() -> None:
    st.session_state["_dirty"] = True


# =========================================================
# STYLE (tu CSS)
# =========================================================
st.markdown(
    """
<style>
html, body, .stApp, .main, .block-container,
div[data-testid="stExpander"], div[data-testid="stTab"],
div[data-testid="stHeader"], .stTabs, [data-testid="stVerticalBlock"] {
    background-color: #ffffff !important;
    color: #000000 !important;
}
.peticiones-title {
    font-size: 2.5rem; font-weight: 800; color: #000000;
    margin-top: 40px; margin-bottom: 20px;
    padding-bottom: 10px; border-bottom: 2px solid #000000; width: 100%;
}
.section-header {
    background: #000; color: #fff; padding: 8px;
    font-weight: bold; margin-top: 20px; margin-bottom: 10px;
}
.table-row {
    border: 1px solid #000000; margin-top: -1px;
    background-color: #ffffff !important; display: flex; align-items: center; width: 100%;
}
.cell-content { padding: 8px 12px; display: flex; flex-direction: column; justify-content: center; }
.stButton>button {
    width: 100% !important; border-radius: 0px !important; font-weight: 700 !important;
    height: 40px; text-transform: uppercase; border: 1px solid #000000 !important; font-size: 0.7rem !important;
}
.stButton>button[kind="secondary"] { background-color: #ffffff !important; color: #000000 !important; }
.stButton>button[kind="primary"] { background-color: #0052FF !important; color: #ffffff !important; border: none !important; }
.summary-box {
    border: 2px solid #000000; padding: 15px; margin-top: 20px;
    background-color: #ffffff !important; font-weight: bold;
    display: flex; justify-content: space-between; color: #000000 !important;
}
@media (max-width: 600px) {
    .peticiones-title { font-size: 1.8rem; margin-top: 20px; }
    .summary-box { flex-direction: column; gap: 5px; }
    .stButton>button { font-size: 0.75rem !important; height: 48px; }
}
</style>
""",
    unsafe_allow_html=True,
)

# =========================================================
# UTILIDADES
# =========================================================
def _norm_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _find_col(df: pd.DataFrame, candidates: list[str]):
    low_map = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in low_map:
            return low_map[cand.lower()]
    for c in df.columns:
        cl = c.lower()
        for cand in candidates:
            if cand.lower() in cl:
                return c
    return None


def _clean_ean(x) -> str:
    if pd.isna(x):
        return ""
    s = str(x).strip()
    if s.lower() == "nan":
        return ""
    if s.endswith(".0"):
        s = s[:-2]
    return s.strip()


def read_excel_any(uploaded_file):
    # xlsx -> openpyxl ; xls -> xlrd (requiere xlrd==2.0.1)
    try:
        return pd.read_excel(uploaded_file, engine="openpyxl")
    except Exception:
        return pd.read_excel(uploaded_file, engine="xlrd")


def make_excel_download(df: pd.DataFrame) -> bytes:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="INCIDENCIAS")
    out.seek(0)
    return out.getvalue()


# Normalización fuerte + mapa básico de tallas
TALLA_MAP = {
    "x-small": "xs", "x small": "xs", "xs": "xs",
    "small": "s", "s": "s",
    "medium": "m", "m": "m",
    "large": "l", "l": "l",
    "x-large": "xl", "x large": "xl", "xl": "xl",
    "xxl": "xxl", "xxxl": "xxxl",
    "1": "1", "2": "2", "3": "3", "4": "4", "5": "5",
}


def norm_txt(x) -> str:
    if x is None:
        return ""
    s = str(x).strip().casefold()
    s = re.sub(r"\s+", " ", s)
    return s


def norm_color(x) -> str:
    s = norm_txt(x)
    s = s.replace(" - ", "-").replace(" / ", "/")
    return s


def norm_talla(x) -> str:
    s = norm_txt(x)
    return TALLA_MAP.get(s, s)


def looks_like_talla(token: str) -> bool:
    """
    Heurística simple: si el token normalizado coincide con una talla conocida,
    o si es del tipo XS/S/M/L/XL/XXL/XXXL, o numérico corto.
    """
    t = norm_talla(token)
    if t in set(TALLA_MAP.values()):
        return True
    if re.fullmatch(r"(xs|s|m|l|xl|xxl|xxxl)", t):
        return True
    if re.fullmatch(r"\d{1,2}", t):
        return True
    return False


def parse_producto_linea(raw: str):
    """
    Admite:
      [262200] Blusa Elyr (Blanco Lagoon, XS)   -> 2 atributos
      [262200] Blusa Elyr (XS)                 -> 1 atributo
      [262200] Blusa Elyr (Blanco Lagoon)      -> 1 atributo
    Devuelve:
      ref_imp, attr1, attr2, num_attrs
    """
    if not isinstance(raw, str):
        return None, None, None, 0
    s = raw.strip()

    m = re.search(r"\[(.*?)\]", s)
    if not m:
        return None, None, None, 0
    ref = m.group(1).strip()

    pm = re.search(r"\((.*)\)\s*$", s)
    if not pm:
        return ref, None, None, 0

    inside = pm.group(1).strip()
    if not inside:
        return ref, None, None, 0

    if "," in inside:
        a1, a2 = inside.rsplit(",", 1)
        return ref, a1.strip(), a2.strip(), 2

    return ref, inside.strip(), None, 1


# =========================================================
# CATÁLOGO
# =========================================================
@st.cache_data
def load_catalogue(path="catalogue.xlsx"):
    if not os.path.exists(path):
        return None, f"No encuentro '{path}' en la carpeta de la app."

    try:
        df = pd.read_excel(path, engine="openpyxl")
        df = _norm_cols(df)

        ean_col = _find_col(df, ["EAN", "Ean", "codigo ean", "código ean", "ean code"])
        if not ean_col:
            return None, f"Catálogo leído, pero NO encuentro columna EAN. Columnas: {list(df.columns)}"

        df["EAN"] = df[ean_col].apply(_clean_ean)

        ref_col = _find_col(df, ["Referencia", "ref", "reference"])
        if ref_col and ref_col != "Referencia":
            df["Referencia"] = df[ref_col].astype(str).str.strip()
        elif "Referencia" in df.columns:
            df["Referencia"] = df["Referencia"].astype(str).str.strip()
        else:
            df["Referencia"] = ""

        for opt in ["Nombre", "Color", "Talla", "Colección", "Categoría", "Familia"]:
            col = _find_col(df, [opt])
            if col and col != opt:
                df[opt] = df[col]
            if opt not in df.columns:
                df[opt] = ""

        # Normalizados
        df["ref_n"] = df["Referencia"].apply(norm_txt)
        df["color_n"] = df["Color"].apply(norm_color)
        df["talla_n"] = df["Talla"].apply(norm_talla)

        df["search_blob"] = (
            df["EAN"].astype(str)
            + " "
            + df["Referencia"].astype(str)
            + " "
            + df["Nombre"].astype(str)
            + " "
            + df["Color"].astype(str)
            + " "
            + df["Talla"].astype(str)
        ).str.lower()

        return df, None
    except Exception as e:
        return None, f"Error leyendo catálogo: {e}"


def build_catalog_indexes(df_cat: pd.DataFrame):
    """
    Construye índices para matching:
      - exact: (ref,color,talla) -> [rows]
      - ref_color: (ref,color) -> [rows]
      - ref_talla: (ref,talla) -> [rows]
      - ref_only: ref -> [rows]
    """
    exact = {}
    ref_color = {}
    ref_talla = {}
    ref_only = {}

    def add(d, k, row):
        d.setdefault(k, []).append(row)

    for _, r in df_cat.iterrows():
        row = {
            "EAN": str(r.get("EAN", "")).strip(),
            "Referencia": r.get("Referencia", ""),
            "Nombre": r.get("Nombre", ""),
            "Color": r.get("Color", ""),
            "Talla": r.get("Talla", ""),
            "ref_n": r.get("ref_n", ""),
            "color_n": r.get("color_n", ""),
            "talla_n": r.get("talla_n", ""),
        }
        add(exact, (row["ref_n"], row["color_n"], row["talla_n"]), row)
        add(ref_color, (row["ref_n"], row["color_n"]), row)
        add(ref_talla, (row["ref_n"], row["talla_n"]), row)
        add(ref_only, row["ref_n"], row)

    return exact, ref_color, ref_talla, ref_only


def pick_unique(rows: list[dict]):
    """
    Devuelve (row, None) si hay un match único; si 0 -> (None,'NO_ENCONTRADO'); si >1 -> (None,'AMBIGUO')
    """
    if not rows:
        return None, "NO_ENCONTRADO"
    if len(rows) == 1:
        return rows[0], None
    return None, "AMBIGUO"


def match_producto(ref_imp: str, a1: str | None, a2: str | None, n_attrs: int,
                   idx_exact, idx_ref_color, idx_ref_talla, idx_ref_only):
    """
    Devuelve: (matched_row | None, reason | None, strategy_text)
    reason:
      - NO_ENCONTRADO
      - AMBIGUO
      - SIN_EAN
    strategy_text:
      - EXACTO ref+color+talla
      - ref+color
      - ref+talla
      - ref+atributo (probado como color/talla)
    """
    ref_n = norm_txt(ref_imp)

    # 2 atributos -> intentamos exacto
    if n_attrs == 2:
        color_n = norm_color(a1)
        talla_n = norm_talla(a2)
        row, err = pick_unique(idx_exact.get((ref_n, color_n, talla_n), []))
        if row:
            if not row.get("EAN"):
                return None, "SIN_EAN", "EXACTO ref+color+talla"
            return row, None, "EXACTO ref+color+talla"
        return None, err, "EXACTO ref+color+talla"

    # 1 atributo -> probar por el atributo que haya
    if n_attrs == 1 and a1:
        token = a1.strip()
        token_is_talla = looks_like_talla(token)

        # Si parece talla -> primero ref+talla
        if token_is_talla:
            talla_n = norm_talla(token)
            row, err = pick_unique(idx_ref_talla.get((ref_n, talla_n), []))
            if row:
                if not row.get("EAN"):
                    return None, "SIN_EAN", "ref+talla"
                return row, None, "ref+talla"
            # si no se encuentra por talla, intentamos ref+color por si venía talla rara
            color_n = norm_color(token)
            row, err2 = pick_unique(idx_ref_color.get((ref_n, color_n), []))
            if row:
                if not row.get("EAN"):
                    return None, "SIN_EAN", "ref+color"
                return row, None, "ref+color"
            # si ambos fallan, devolvemos el más informativo: si alguno fue AMBIGUO, avisarlo
            if err == "AMBIGUO" or err2 == "AMBIGUO":
                return None, "AMBIGUO", "ref+(atributo)"
            return None, "NO_ENCONTRADO", "ref+(atributo)"

        # Si NO parece talla -> primero ref+color
        color_n = norm_color(token)
        row, err = pick_unique(idx_ref_color.get((ref_n, color_n), []))
        if row:
            if not row.get("EAN"):
                return None, "SIN_EAN", "ref+color"
            return row, None, "ref+color"

        # si no, intentar ref+talla por si el color era en realidad talla
        talla_n = norm_talla(token)
        row, err2 = pick_unique(idx_ref_talla.get((ref_n, talla_n), []))
        if row:
            if not row.get("EAN"):
                return None, "SIN_EAN", "ref+talla"
            return row, None, "ref+talla"

        if err == "AMBIGUO" or err2 == "AMBIGUO":
            return None, "AMBIGUO", "ref+(atributo)"
        return None, "NO_ENCONTRADO", "ref+(atributo)"

    return None, "NO_ENCONTRADO", "sin atributos"


# =========================================================
# SESSION STATE INIT
# =========================================================
st.session_state.setdefault("carrito", {})
st.session_state.setdefault("search_key", 0)
st.session_state.setdefault("_dirty", False)
st.session_state.setdefault("_hydrated", False)
st.session_state.setdefault("origen", "PET Almacén Badalona")
st.session_state.setdefault("destino", "PET T002 Marbella")
st.session_state.setdefault("ref_peticion", "")
st.session_state.setdefault("fecha_str", datetime.now().strftime("%Y-%m-%d"))

# =========================================================
# HYDRATE ONCE
# =========================================================
if not st.session_state._hydrated:
    val = ls_get(LS_KEY, "__ls_payload")
    if val:
        try:
            _apply_state(json.loads(val))
        except Exception:
            pass
    st.session_state._hydrated = True


# =========================================================
# UI
# =========================================================
st.markdown('<div class="peticiones-title">Peticiones</div>', unsafe_allow_html=True)

df_cat, cat_err = load_catalogue("catalogue.xlsx")
if cat_err:
    st.error(cat_err)
    st.stop()

idx_exact, idx_ref_color, idx_ref_talla, idx_ref_only = build_catalog_indexes(df_cat)

# ---------------------------------------------------------
# 1) CABECERA
# ---------------------------------------------------------
c1, c2, c3 = st.columns(3)

try:
    fecha_default = datetime.strptime(st.session_state.fecha_str, "%Y-%m-%d")
except Exception:
    fecha_default = datetime.now()
    st.session_state.fecha_str = fecha_default.strftime("%Y-%m-%d")

fecha = c1.date_input("FECHA", fecha_default, key="fecha_widget", on_change=mark_dirty)
st.session_state.fecha_str = fecha.strftime("%Y-%m-%d")

c2.selectbox("ORIGEN", ["PET Almacén Badalona", "ALM-CENTRAL"], key="origen", on_change=mark_dirty)
c3.selectbox("DESTINO", ["PET T002 Marbella", "ALM-TIENDA"], key="destino", on_change=mark_dirty)
st.text_input("REFERENCIA PETICIÓN", key="ref_peticion", on_change=mark_dirty)

fecha_str = st.session_state.fecha_str
origen = st.session_state.origen
destino = st.session_state.destino
ref_peticion = st.session_state.ref_peticion

st.write("---")

# ---------------------------------------------------------
# 2) IMPORTADOR MASIVO
#    - A = producto
#    - cantidad en C si hay algún valor >0, si no en B
#    - cruza por:
#         2 attrs: ref+color+talla
#         1 attr: ref + (color o talla)
#    - si cruza pero NO tiene EAN -> aviso específico
# ---------------------------------------------------------
st.markdown('<div class="section-header">📂 IMPORTACIÓN DE VENTAS / REPOSICIÓN</div>', unsafe_allow_html=True)

u1, u2 = st.columns([3, 1])

archivo_v = u1.file_uploader(
    "Sube Excel TPV (A=Producto, Cantidad en B o C) - xlsx/xls",
    type=["xlsx", "xls"],
    key="upload_excel",
    label_visibility="visible",
)

if archivo_v is not None:
    st.info(f"Archivo cargado: **{archivo_v.name}**")

inject = u2.button("INJECTAR AL CARRITO", type="primary", disabled=(archivo_v is None))

if inject:
    try:
        df_v = read_excel_any(archivo_v)

        if df_v.shape[1] < 2:
            st.error("El Excel debe tener al menos 2 columnas: A=Producto y B=Cantidad (o C=Cantidad).")
            st.stop()

        prod_series = df_v.iloc[:, 0].astype(str)

        qty_b = pd.to_numeric(df_v.iloc[:, 1], errors="coerce")
        qty_c = None
        use_c = False
        if df_v.shape[1] >= 3:
            qty_c = pd.to_numeric(df_v.iloc[:, 2], errors="coerce")
            if qty_c.notna().any() and (qty_c.fillna(0) > 0).any():
                use_c = True

        qty_series = (qty_c if use_c else qty_b).fillna(0).astype(int)

        work = pd.DataFrame({"prod_raw": prod_series, "qty": qty_series})
        work["prod_raw"] = work["prod_raw"].astype(str).str.strip()
        work = work[work["qty"] > 0].copy()

        # Solo filas con referencia entre corchetes y paréntesis al final
        mask = work["prod_raw"].str.contains(r"\[.*?\].*\(.*\)", regex=True, na=False)
        work = work[mask].copy()

        if work.empty:
            st.session_state["import_result"] = {
                "added": 0,
                "no_match": 0,
                "no_ean": 0,
                "ambiguous": 0,
                "used_qty_col": "C" if use_c else "B",
                "note": "No se encontraron filas válidas (patrón [REF] ... (atributos) y qty>0).",
            }
            st.session_state["import_issues_records"] = []
            st.rerun()

        # Parse
        parsed = work["prod_raw"].apply(parse_producto_linea)
        work["ref_imp"] = parsed.apply(lambda t: t[0])
        work["a1"] = parsed.apply(lambda t: t[1])
        work["a2"] = parsed.apply(lambda t: t[2])
        work["n_attrs"] = parsed.apply(lambda t: t[3])

        # Agrupar duplicados por lo parseado (reduce ruido)
        grouped = (
            work.groupby(["ref_imp", "a1", "a2", "n_attrs"], as_index=False)["qty"]
            .sum()
            .copy()
        )

        added_lines = 0
        issues = []
        c_no_match = 0
        c_no_ean = 0
        c_amb = 0

        for _, r in grouped.iterrows():
            ref_imp = r["ref_imp"]
            a1 = r["a1"]
            a2 = r["a2"]
            n_attrs = int(r["n_attrs"])
            qty = int(r["qty"])

            row, reason, strategy = match_producto(
                ref_imp, a1, a2, n_attrs,
                idx_exact, idx_ref_color, idx_ref_talla, idx_ref_only
            )

            if row is None:
                if reason == "SIN_EAN":
                    c_no_ean += 1
                    motivo = "Referencia encontrada pero SIN EAN en catálogo"
                elif reason == "AMBIGUO":
                    c_amb += 1
                    motivo = "Cruce AMBIGUO (varios productos posibles). Añade más detalle (color/talla)."
                else:
                    c_no_match += 1
                    motivo = "No se ha encontrado en catálogo con los datos disponibles"

                issues.append(
                    {
                        "Motivo": motivo,
                        "Estrategia": strategy,
                        "Producto_raw": f"[{ref_imp}] ({a1}{', ' + a2 if a2 else ''})",
                        "Referencia": ref_imp,
                        "Atributo_1": a1,
                        "Atributo_2": a2,
                        "Cantidad": qty,
                    }
                )
                continue

            # Añadir al carrito
            ean = str(row["EAN"]).strip()
            if ean in st.session_state.carrito:
                st.session_state.carrito[ean]["Cantidad"] += qty
            else:
                st.session_state.carrito[ean] = {
                    "Ref": row.get("Referencia", ""),
                    "Nom": row.get("Nombre", ""),
                    "Col": row.get("Color", "-"),
                    "Tal": row.get("Talla", "-"),
                    "Cantidad": qty,
                }
            added_lines += 1

        st.session_state["import_result"] = {
            "added": int(added_lines),
            "no_match": int(c_no_match),
            "no_ean": int(c_no_ean),
            "ambiguous": int(c_amb),
            "used_qty_col": "C" if use_c else "B",
            "note": None,
        }
        st.session_state["import_issues_records"] = issues

        mark_dirty()
        st.rerun()

    except ImportError:
        st.error("Para leer .xls necesitas `xlrd==2.0.1` en requirements.txt (o convierte a .xlsx).")
    except Exception as e:
        st.error(f"No he podido importar el Excel: {e}")

# Panel persistente de resultado de importación
if "import_result" in st.session_state and st.session_state["import_result"]:
    res = st.session_state["import_result"]
    added = int(res.get("added", 0))
    no_match = int(res.get("no_match", 0))
    no_ean = int(res.get("no_ean", 0))
    ambiguous = int(res.get("ambiguous", 0))
    used_col = res.get("used_qty_col", "")
    note = res.get("note", None)

    if note:
        st.info(note)

    if added > 0:
        st.success(f"Importación OK ✅ Líneas añadidas/actualizadas: {added} | Cantidad tomada de columna: {used_col}")
    else:
        st.warning(f"No se han añadido líneas | Cantidad tomada de columna: {used_col}")

    if no_match + no_ean + ambiguous > 0:
        msg_parts = []
        if no_match > 0:
            msg_parts.append(f"{no_match} no encontradas")
        if ambiguous > 0:
            msg_parts.append(f"{ambiguous} ambiguas")
        if no_ean > 0:
            msg_parts.append(f"{no_ean} sin EAN en catálogo")
        st.warning("⚠️ Incidencias: " + " | ".join(msg_parts))

        records = st.session_state.get("import_issues_records", [])
        df_issues = pd.DataFrame(records) if records else pd.DataFrame(
            columns=["Motivo", "Estrategia", "Producto_raw", "Referencia", "Atributo_1", "Atributo_2", "Cantidad"]
        )

        with st.expander("Ver incidencias de importación", expanded=False):
            st.dataframe(df_issues, use_container_width=True)

        st.download_button(
            "⬇️ Descargar incidencias (Excel)",
            data=make_excel_download(df_issues),
            file_name="incidencias_importacion.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    if st.button("Ocultar resultado de importación"):
        st.session_state.pop("import_result", None)
        st.session_state.pop("import_issues_records", None)
        st.rerun()

st.write("---")

#
---------------------------------------------------------
# 3) BUSCADOR
# ---------------------------------------------------------
st.markdown('<div class="section-header">🔍 BUSCADOR MANUAL</div>', unsafe_allow_html=True)
f1, f2 = st.columns([2, 1])
busq_txt = f1.text_input("Buscar referencia, nombre o EAN...", key=f"busq_{st.session_state.search_key}")
limite = f2.selectbox("Ver resultados:", [10, 25, 50, 100, 500], index=1, key=f"lim_{st.session_state.search_key}")

filtros_activos = {}
columnas_posibles = ["Colección", "Categoría", "Familia"]
columnas_reales = [c for c in columnas_posibles if c in df_cat.columns]

if columnas_reales:
    cols_f = st.columns(len(columnas_reales))
    for i, col in enumerate(columnas_reales):
        opciones = ["TODOS"] + sorted(
            [x for x in df_cat[col].dropna().astype(str).unique().tolist() if x.strip() != ""]
        )
        filtros_activos[col] = cols_f[i].selectbox(f"{col}", opciones, key=f"f_{col}_{st.session_state.search_key}")

df_res = df_cat
needle = (busq_txt or "").strip().lower()
if needle:
    df_res = df_res[df_res["search_blob"].str.contains(needle, na=False)]

for col, val in filtros_activos.items():
    if val != "TODOS":
        df_res = df_res[df_res[col].astype(str) == str(val)]

if needle or any(v != "TODOS" for v in filtros_activos.values()):
    st.markdown(
        f"<div style='background: #000; color: #fff; padding: 4px; font-size: 0.7rem; text-align: center;'>{len(df_res)} COINCIDENCIAS</div>",
        unsafe_allow_html=True,
    )

    for _, f in df_res.head(limite).iterrows():
        ean = str(f["EAN"]).strip()
        en_car = ean in st.session_state.carrito

        st.markdown('<div class="table-row">', unsafe_allow_html=True)
        c1r, c2r = st.columns([3, 1.5])
        with c1r:
            st.markdown(
                f"<div class='cell-content'><strong>{f.get('Referencia','')}</strong><br>"
                f"<small>{f.get('Nombre','')} ({f.get('Color','-')} / {f.get('Talla','-')})</small></div>",
                unsafe_allow_html=True,
            )
        with c2r:
            label = f"OK ({st.session_state.carrito[ean]['Cantidad']})" if en_car else "AÑADIR"
            if st.button(label, key=f"b_{ean}", type="primary" if en_car else "secondary"):
                if en_car:
                    st.session_state.carrito[ean]["Cantidad"] += 1
                else:
                    st.session_state.carrito[ean] = {
                        "Ref": f.get("Referencia", ""),
                        "Nom": f.get("Nombre", ""),
                        "Col": f.get("Color", "-"),
                        "Tal": f.get("Talla", "-"),
                        "Cantidad": 1,
                    }
                mark_dirty()
                st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

# ---------------------------------------------------------
# 4) LISTA FINAL Y GENERACIÓN
# ---------------------------------------------------------
if st.session_state.carrito:
    st.write("---")
    st.markdown('<div class="section-header">📋 LISTA DE REPOSICIÓN</div>', unsafe_allow_html=True)

    for ean, item in list(st.session_state.carrito.items()):
        st.markdown('<div class="table-row">', unsafe_allow_html=True)
        ca, cb, cc = st.columns([2.5, 1.2, 0.8])

        with ca:
            st.markdown(
                f"<div class='cell-content'><strong>{item.get('Ref','')}</strong><br><small>{item.get('Nom','')}</small></div>",
                unsafe_allow_html=True,
            )

        with cb:
            new_qty = st.number_input(
                "C",
                1,
                9999,
                int(item.get("Cantidad", 1)),
                key=f"q_{ean}",
                label_visibility="collapsed",
                on_change=mark_dirty,
            )
            item["Cantidad"] = int(new_qty)

        with cc:
            if st.button("✕", key=f"d_{ean}"):
                del st.session_state.carrito[ean]
                mark_dirty()
                st.rerun()

        st.markdown("</div>", unsafe_allow_html=True)

    uds = sum(int(it.get("Cantidad", 0)) for it in st.session_state.carrito.values())
    st.markdown(
        f'<div class="summary-box"><div>PIEZAS: {uds}</div><div>MODELOS: {len(st.session_state.carrito)}</div><div>DESTINO: {st.session_state.destino}</div></div>',
        unsafe_allow_html=True,
    )

    cv, cg = st.columns([1, 2])

    if cv.button("LIMPIAR TODO"):
        st.session_state.carrito = {}
        st.session_state.search_key += 1
        mark_dirty()
        st.rerun()

    if cv.button("🧹 OLVIDAR EN ESTE NAVEGADOR"):
        ls_set(LS_KEY, json.dumps({"carrito": {}, "ref_peticion": ""}))
        st.session_state.carrito = {}
        st.session_state.ref_peticion = ""
        st.session_state._dirty = False
        st.rerun()

    if cg.button("GENERAR Y DESCARGAR EXCEL", type="primary"):
        try:
            if os.path.exists("peticion.xlsx"):
                with open("peticion.xlsx", "rb") as f:
                    tpl_bytes = f.read()
                wb = load_workbook(io.BytesIO(tpl_bytes))
                ws = wb.active
            else:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.append(["Fecha", "Origen", "Destino", "Referencia", "EAN", "Cantidad"])

            for ean, it in st.session_state.carrito.items():
                ws.append([fecha_str, origen, destino, ref_peticion, ean, int(it["Cantidad"])])

            out = io.BytesIO()
            wb.save(out)
            out.seek(0)

            st.download_button(
                "📥 GUARDAR ARCHIVO REPO",
                out.getvalue(),
                file_name=f"REPO_{destino}.xlsx",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"No he podido generar el Excel: {e}")

# =========================================================
# SAVE TO LOCAL STORAGE IF DIRTY
# =========================================================
if st.session_state._dirty:
    try:
        payload = _serialize_state()
        ls_set(LS_KEY, json.dumps(payload))
        st.session_state._dirty = False
    except Exception:
        st.session_state._dirty = False
